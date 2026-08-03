"""PHARMAC (New Zealand) Pharmaceutical Schedule scraper — migrated onto the
engine.

Downloads the Community Medicines sheet from the PHARMAC CPSReporting xlsx,
yielding subsidised generic/brand medicine names. Single bulk XLSX download,
cached to ``{output_dir}/_pharmac_drugs.xlsx`` (exactly like the original),
then parsed and returned in one shot with ``next_cursor=None``.

The original deduplicated on the composite key ``(chemical, presentation,
brand)`` rather than a single column, so ``id_field`` is left empty and the
same composite-key dedup is reproduced manually inside :meth:`fetch` (the
whole dataset is one page anyway, so this preserves identical output).
DEVIATION: resume-granularity coarsens from "per-row" (original's in-memory
`seen` set, persisted only via the completed whole-file checkpoint) to
"whole dataset" — same as the original in practice, since it was always a
single-shot crawl with no partial resume either.

Schema per row:
  chemical, presentation, brand, pharmacode, nzmt_ctpp_id,
  subsidy, fully_subsidised, language

Run it::

    python -m metadatarr.scrapers pharmac_drugs [--output DIR]
"""
from __future__ import annotations

from typing import List

from metadatarr.scrapers.engine import Page, Source, register, run_cli

XLSX_URL = "https://schedule.pharmac.govt.nz/latest/CPSReporting.xlsx"
SHEET_NAME = "Community Medicines"
DATA_START = 3  # 0-indexed first data row


def _parse_workbook(ws) -> List[dict]:
    seen: set = set()
    rows_out: List[dict] = []
    row_idx = 0
    for row in ws.iter_rows(values_only=True):
        if row_idx < DATA_START:
            row_idx += 1
            continue
        row_idx += 1

        # Col A: Chemical, B: Presentation, C: Brand, D: Pharmacode,
        # E: NZMT CTPP ID, F: Price, G: Subsidy, H: Alternate,
        # I: Per, J: Fully subsidised
        chemical = str(row[0]).strip() if row[0] is not None else ""
        if not chemical or chemical.lower() == "none":
            continue

        presentation = str(row[1]).strip() if row[1] is not None else ""
        brand = str(row[2]).strip() if row[2] is not None else ""
        pharmacode = str(row[3]).strip() if row[3] is not None else ""
        nzmt_ctpp_id = str(row[4]).strip() if row[4] is not None else ""
        subsidy = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
        fully_subsidised = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ""

        dedup_key = (chemical, presentation, brand)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        rows_out.append({
            "chemical": chemical,
            "presentation": presentation,
            "brand": brand,
            "pharmacode": pharmacode,
            "nzmt_ctpp_id": nzmt_ctpp_id,
            "subsidy": subsidy,
            "fully_subsidised": fully_subsidised,
            "language": "en-NZ",
        })
    return rows_out


@register
class PharmacDrugsSource(Source):
    name = "pharmac_drugs"
    id_field = ""
    default_delay = 0.0

    def initial_cursor(self) -> int:
        return 0

    def fetch(self, cursor: int) -> Page:
        if cursor is None:
            return [], None

        import openpyxl
        import requests

        output_dir = self._output_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        cache_xlsx = output_dir / "_pharmac_drugs.xlsx"

        if not cache_xlsx.exists():
            s = requests.Session()
            s.headers["User-Agent"] = (
                "Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0"
            )
            r = s.get(XLSX_URL, timeout=120, stream=True)
            r.raise_for_status()
            data = b""
            for chunk in r.iter_content(65536):
                data += chunk
            cache_xlsx.write_bytes(data)

        wb = openpyxl.load_workbook(cache_xlsx, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            available = wb.sheetnames
            wb.close()
            raise ValueError(f"Sheet '{SHEET_NAME}' not found; available: {available}")
        ws = wb[SHEET_NAME]
        rows = _parse_workbook(ws)
        wb.close()

        return rows, None


if __name__ == "__main__":
    raise SystemExit(run_cli(PharmacDrugsSource))
